#!/usr/bin/env python3

# Author: Ajay Singh
# Date  : 07-Jul-2022
# Purpose: Compare 2 excel files data. Data from 1 (old) is searched in latest (2) and if,
#          there is any change in it, new data is copied to old file in an extra column.
# Algorithm:
"""
Create a log file for writing
Open latest file (2-*.xlsx) for reading
Open old file (1-*.xlsx) for reading & writing row-by-row. This file will also be updated with modified data.
    - Add 3 columns to update the old excel
        - 'PID Not Found'
        - 'Updated Description'
        - 'Updated Price'
    - Get cell data of columns:
        - 'Product', 
        - 'Product Description'
        - 'Price in USD'
    - Search in 2-*.xlsx for 'Product' string in 'Product' column
        - if NOT-FOUND:
            - Update Log file: PID: NOT-FOUND: 'Product' string
            - update 'NOT-FOUND' in 'PID Not Found' column in 1-*.xlsx
        - if FOUND:
            - Update Log file: PID: FOUND: 'Product' string
            - Match 'Product Description' column string of this row with that from 1-*.xlsx
                - if MATCHED: 
                    - Update Log file: PID: Description matched: 'Product Description' string
                - else: 
                    - Update Log file: PID: Description NOT matched: Description OLD: Description NEW
                    - Update 1-*.xlsx column 'Updated Description' with latest 'Product Description' as per 2-*.xlsx
                    
            - Match 'Price in USD' column string of this row with that from 1-*.xlsx
                - if MATCHED: 
                    - Update Log file: PID: Price matched: 'Price in USD' string
                - else: 
                    - Update Log file: PID: Price NOT matched: Price OLD: Price NEW
                    - Update 1-*.xlsx column 'Updated Price' with latest 'Price in USD' as per 2-*.xlsx
Close all the files
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os
from time import sleep

# Create a log file
log_filename = "log_run.txt"
f_logfile = open(log_filename, "w")
def xprint(str):
    try:
        f_logfile.write('%s\n' % str)
        print('%s' % str)
    except:
        print('*** ERROR: Failed to write : %s ***' % str)

# Method of action
def update_excel(old_excel, latest_excel):
    try:
        xprint("old_excel = %s" % old_excel)
        xprint("latest_excel = %s" % latest_excel)
        # Sheet name will be default "Sheet1"
        latest_sheet_name = old_sheet_name = "Sheet1"
        
        # Open the old excel for writing (1-*.xlsx)
        try:
            wb_old_excel = load_workbook(old_excel)
            ws1 = wb_old_excel[old_sheet_name]
        except Exception as ex:
            template = "EXCEPTION:: An exception of type {0} occurred. Arguments:\n{1!r}"
            xprint(template.format(type(ex).__name__, ex.args))

        # Open the latest excel for reading (2-*.xlsx)
        try:
            wb_latest_excel = load_workbook(latest_excel)
            ws2 = wb_latest_excel[latest_sheet_name]
        except Exception as ex:
            template = "EXCEPTION:: An exception of type {0} occurred. Arguments:\n{1!r}"
            xprint(template.format(type(ex).__name__, ex.args))
        
        # Add 3 columns to update the old excel
        # 'PID Not Found'
        # 'Updated Description'
        # 'Updated Price'
        columns_to_be_added = [
            "PID Not Found",
            "Updated Description",
            "Updated Price",
        ]
        xprint("Columns count = %s" % ws1.max_column)
        #max_col = ws1.max_column
        max_col = 4 # Hard-coded to avoid creating recurring columns on rerun
        for col_value in columns_to_be_added:
            ws1.cell(row=1, column=max_col+1, value=col_value)
            max_col += 1
        # Save and close
        wb_old_excel.save(old_excel)
        sleep(1)
        # Re-Open the old excel again for writing (1-*.xlsx)
        try:
            wb_old_excel = load_workbook(old_excel)
            ws1 = wb_old_excel[old_sheet_name]
        except Exception as ex:
            template = "EXCEPTION:: An exception of type {0} occurred. Arguments:\n{1!r}"
            xprint(template.format(type(ex).__name__, ex.args))
        sleep(1)
        xprint("Updated sheet with the needed columns")
        xprint("Columns count = %s" % ws1.max_column)

        xprint("ws1: max_row is in row %s and max_column is in column %s" % (ws1.max_row, ws1.max_column))
        xprint("ws2: max_row is in row %s and max_column is in column %s" % (ws2.max_row, ws2.max_column))
        
        # Get all the column names in both the sheets
        # Create a dictionary of column names
        ws1_col_names = {}
        ws2_col_names = {}
        current  = 0
        for col in ws1.iter_cols(1, ws1.max_column):
            ws1_col_names[col[0].value] = current
            current += 1
        current = 0
        for col in ws2.iter_cols(1, ws2.max_column):
            ws2_col_names[col[0].value] = current
            current += 1
        
        #
        # Now you can access by column name
        # 'Product'
        # 'Product Description'
        # 'Price in USD'
        #
        # 'PID Not Found'
        # 'Updated Description'
        # 'Updated Price'
        #
        #for row_cells in ws1.iter_rows(min_row=2, max_row=5):
        for row_cells in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            ws1_product = row_cells[ws1_col_names['Product']].value
            ws1_description = row_cells[ws1_col_names['Product Description']].value
            ws1_price = row_cells[ws1_col_names['Price in USD']].value
            # Search in ws2 now
            b_pid_found = False
            b_description_found = False
            b_price_found = False
            for row_cells_2 in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                if ws1_product == row_cells_2[ws2_col_names['Product']].value:
                    b_pid_found = True
                    xprint("PID: FOUND: %s" % ws1_product)
                    ws2_description = row_cells_2[ws2_col_names['Product Description']].value
                    ws2_price_temp = row_cells_2[ws2_col_names['Price in USD']].value
                    ws2_price = ws2_price_temp.replace('$', '')
                    if ws1_description == ws2_description:
                        b_description_found = True
                        xprint("PID: Description matched: %s" % ws1_description)
                    else:
                        xprint("PID: Description NOT matched:")
                        row_cells[ws1_col_names['Updated Description']].value = ws2_description
                        xprint("Expected: %s" % ws1_description)
                        xprint("Found   : %s" % ws2_description)
                    if float(ws1_price) == float(ws2_price):
                        b_price_found = True
                        xprint("PID: Price matched: %s" % ws1_price)
                    else:
                        xprint("PID: Price NOT matched:")
                        row_cells[ws1_col_names['Updated Price']].value = ws2_price
                        xprint("Expected: %s" % ws1_price)
                        xprint("Found   : %s" % ws2_price)
            if not b_pid_found:
                row_cells[ws1_col_names['PID Not Found']].value = 'NOT-FOUND'
                xprint("PID: NOT-FOUND: %s" % ws1_product)
                
            xprint("")
        xprint("-"*50)
        
        # Save the wb
        wb_old_excel.save(old_excel)
        wb_latest_excel.save(latest_excel)

    except:
        print("Exception in update_excel")
        sys.exit(1)


if __name__ == "__main__":
    try:
        old_excel = "1-Report_old.xlsx"
        latest_excel = "2-Report_new.xlsx"

        if ".xlsx" not in latest_excel or ".xlsx" not in old_excel:
            xprint("Please specify files with .xlsx extension")
            raise()
        update_excel(old_excel, latest_excel)
    except:
        print("Error: " + os.path.basename(__file__) + " <latest_excel.xlsx> <old_excel.xlsx>")
        sys.exit(1)



